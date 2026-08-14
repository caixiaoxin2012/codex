from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


TIA_COLUMNS = [
    "Name", "Path", "Data Type", "Logical Address", "Comment",
    "Hmi Visible", "Hmi Accessible", "Hmi Writable",
]
TRACE_COLUMNS = ["项目", "页码", "源文件", "页标题"]
REFERENCE_COLUMNS = TIA_COLUMNS + TRACE_COLUMNS

TIA_BLUE = "2F5597"
TITLE_BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
RAW_BROWN = "7F6000"
STEP_GREEN = "548235"


def _text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _clean_comment(value: object) -> str:
    """Remove extraction noise while keeping the most descriptive nearby text."""
    text = re.sub(r"[\x00-\x1f\x7f]", " ", _text(value))
    parts = []
    for part in re.split(r"\s*\|\s*", text):
        clean = re.sub(r"\s+", " ", part).strip(" -:;|,，。")
        if clean and clean not in parts:
            parts.append(clean)
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0][:240]
    # PDF association can append a neighboring wiring line. Prefer concise Chinese text.
    def score(part: str) -> tuple[int, int, int]:
        chinese = len(re.findall(r"[\u4e00-\u9fff]", part))
        return (1 if chinese else 0, chinese, -len(part))
    return max(parts, key=score)[:240]


def infer_project_name(source_path: Path | None) -> str:
    """Infer a short, TIA-safe project prefix from the input file name."""
    stem = source_path.stem if source_path else "PROJECT"
    match = re.match(r"[A-Za-z0-9]+(?:[-_][A-Za-z0-9]+)*", stem)
    candidate = match.group(0) if match else stem
    candidate = re.split(r"_(?:PLC|IO|EPLAN)(?:_|$)", candidate, maxsplit=1, flags=re.IGNORECASE)[0]
    candidate = re.sub(r"[^A-Za-z0-9_]", "_", candidate).strip("_")
    if not candidate:
        candidate = "PROJECT"
    if candidate[0].isdigit():
        candidate = f"P_{candidate}"
    return candidate[:40]


def _tia_logical_address(value: object) -> str:
    address = _text(value).upper().replace(" ", "")
    if not address or address.startswith("%"):
        return address
    if re.fullmatch(
        r"(?:I|Q|M)\d+\.\d+|(?:PIW|PQW|AIW|AQW|IW|QW|ID|QD|IL|QL|MB|MW|MD|ML)\d+|"
        r"DB\d+\.DB[XBWD]\d+(?:\.\d+)?",
        address,
    ):
        return f"%{address}"
    return address


def _is_tia_importable(logical: str, data_type: str) -> bool:
    if not logical or not data_type:
        return False
    return bool(re.fullmatch(
        r"%(?:[IQM]\d+\.\d+|(?:PIW|PQW|AIW|AQW|IW|QW|ID|QD|IL|QL|MB|MW|MD|ML)\d+|"
        r"DB\d+\.DB[XBWD]\d+(?:\.\d+)?)",
        logical,
    ))


def _address_sort_key(logical: str) -> tuple[int, int, int, int, str]:
    value = logical.lstrip("%")
    match = re.fullmatch(r"([IQM])([BWDL]?)(\d+)(?:\.(\d+))?", value)
    if match:
        area, size, byte, bit = match.groups()
        return ({"I": 0, "Q": 1, "M": 2}[area], int(byte), int(bit or -1), {"": 0, "B": 1, "W": 2, "D": 3, "L": 4}[size], value)
    db = re.fullmatch(r"DB(\d+)\.DB([XBWD])(\d+)(?:\.(\d+))?", value)
    if db:
        block, size, byte, bit = db.groups()
        return (3, int(block), int(byte), int(bit or -1), size)
    return (9, 0, 0, 0, value)


def _tia_data_type(io_type: object, address: object) -> str:
    kind = _text(io_type)
    logical = _tia_logical_address(address)
    if kind in {"DI", "DO"} or re.fullmatch(r"%[IQM]\d+\.\d+", logical):
        return "Bool"
    if kind in {"AI", "AO"}:
        return "DInt" if re.match(r"%(?:I|Q)[DL]", logical) else "Int"
    if kind == "Memory":
        if re.match(r"%M\d+\.\d+", logical):
            return "Bool"
        if logical.startswith("%MB"):
            return "Byte"
        if logical.startswith("%MW"):
            return "Word"
        if logical.startswith(("%MD", "%ML")):
            return "DWord"
    if kind == "DB":
        for token, data_type in ((".DBX", "Bool"), (".DBB", "Byte"), (".DBW", "Word"), (".DBD", "DWord")):
            if token in logical:
                return data_type
    return ""


def _tag_name(project: str, logical_address: str) -> str:
    address = re.sub(r"[^A-Za-z0-9_]", "_", logical_address.lstrip("%")).strip("_")
    return f"{project}_{address or 'TAG'}"[:120]


def build_tia_table(
    detail: pd.DataFrame,
    source_path: Path | None = None,
    project_name: str | None = None,
) -> pd.DataFrame:
    """Create rows that mirror the supplied TIA workbook's 12-column layout."""
    project = project_name or infer_project_name(source_path)
    source_name = source_path.name if source_path else ""
    candidates: dict[str, tuple[tuple[int, int, int, int], dict[str, object]]] = {}
    confidence_rank = {"high": 3, "medium": 2, "low": 1}
    for source_index, row in detail.iterrows():
        logical = _tia_logical_address(row.get("标准地址", ""))
        data_type = _tia_data_type(row.get("类型", ""), logical)
        if not _is_tia_importable(logical, data_type):
            continue
        comment = _clean_comment(row.get("说明", ""))
        page = row.get("页码", "")
        record = {
            "Name": _tag_name(project, logical),
            "Path": "",
            "Data Type": data_type,
            "Logical Address": logical,
            "Comment": comment,
            "Hmi Visible": "True",
            "Hmi Accessible": "True",
            "Hmi Writable": "True",
            "项目": project,
            "页码": "" if pd.isna(page) else page,
            "源文件": source_name,
            "页标题": _text(row.get("页标题", "")),
        }
        confidence = confidence_rank.get(_text(row.get("关联置信度", "")).lower(), 0)
        has_chinese = bool(re.search(r"[\u4e00-\u9fff]", comment))
        quality = (1 if comment else 0, confidence, 1 if has_chinese else 0, -int(source_index))
        previous = candidates.get(logical)
        if previous is None or quality > previous[0]:
            candidates[logical] = (quality, record)
    records = [item[1] for item in candidates.values()]
    records.sort(key=lambda record: _address_sort_key(str(record["Logical Address"])))
    return pd.DataFrame(records, columns=REFERENCE_COLUMNS)


def export_csv(detail: pd.DataFrame, output_path: Path) -> Path:
    path = output_path.with_suffix(".csv")
    detail.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def export_tia_csv(detail: pd.DataFrame, output_path: Path, source_path: Path | None = None) -> Path:
    """Export the first eight TIA columns as a UTF-8 BOM CSV."""
    path = output_path.with_name(f"{output_path.stem}_TIA.csv")
    build_tia_table(detail, source_path)[TIA_COLUMNS].to_csv(path, index=False, encoding="utf-8-sig")
    return path


def _append_frame(worksheet, frame: pd.DataFrame) -> None:
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(list(row))


def _style_header(worksheet, color: str, max_column: int) -> None:
    fill = PatternFill("solid", fgColor=color)
    border = Border(bottom=Side(style="thin", color="FFFFFF"))
    for cell in worksheet[1][:max_column]:
        cell.fill = fill
        cell.font = Font(name="Carlito", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _add_table(worksheet, display_name: str, max_column: int, max_row: int) -> None:
    if max_row < 2:
        return
    ref = f"A1:{get_column_letter(max_column)}{max_row}"
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    worksheet.add_table(table)


def _make_tia_sheet(workbook: Workbook, title: str, frame: pd.DataFrame, table_name: str) -> None:
    worksheet = workbook.create_sheet(title)
    _append_frame(worksheet, frame)
    for index, width in enumerate([20, 10, 14, 18, 42, 15, 16, 15, 14, 8, 30, 24], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    _style_header(worksheet, TIA_BLUE, len(REFERENCE_COLUMNS))
    worksheet.freeze_panes = "A2"
    for row in worksheet.iter_rows(min_row=2, max_col=len(REFERENCE_COLUMNS)):
        for cell in row:
            cell.font = Font(name="Carlito", size=11)
            cell.alignment = Alignment(vertical="center")
    _add_table(worksheet, table_name, len(REFERENCE_COLUMNS), worksheet.max_row)


def _make_summary_sheet(workbook: Workbook, frame: pd.DataFrame, project: str, import_sheet: str) -> None:
    worksheet = workbook.active
    worksheet.title = "说明_汇总"
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = "PLC变量表（TIA可导入格式）"
    worksheet["A1"].fill = PatternFill("solid", fgColor=TITLE_BLUE)
    worksheet["A1"].font = Font(name="Carlito", size=16, bold=True, color="FFFFFF")
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells("A2:H2")
    worksheet["A2"] = (
        f"生成时间：{datetime.now():%Y-%m-%d %H:%M}；由 EPLAN Tag Exporter 自动生成，"
        "导入前请复核变量名、数据类型与逻辑地址。"
    )
    for cell in worksheet[3]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(name="Carlito", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center")
    counts = frame["Logical Address"].astype(str)
    di = int(counts.str.match(r"%I\d+\.\d+$").sum())
    do = int(counts.str.match(r"%Q\d+\.\d+$").sum())
    worksheet.append(["项目", "DI数量", "DO数量", "合计", "导入工作表", "建议", "源图纸数量", "备注"])
    worksheet.append([
        project, di, do, len(frame), import_sheet,
        "在 TIA Portal 中导入前先备份项目，并检查地址冲突。", 1,
        "主表已按地址自然排序、去除重复并隔离无效地址；HMI 三项权限默认为 True。",
    ])
    for index, width in enumerate([76.88, 5.88, 6.63, 4.25, 11.38, 20.38, 9.75, 38], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=4, max_row=5, max_col=8):
        for cell in row:
            cell.font = Font(name="Carlito", size=11, bold=cell.row == 4)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _make_raw_sheet(
    workbook: Workbook,
    detail: pd.DataFrame,
    tia_frame: pd.DataFrame,
    project: str,
    source_name: str,
) -> None:
    rows: list[dict[str, object]] = []
    included = set(tia_frame["Logical Address"].astype(str))
    logical_values = [_tia_logical_address(value) for value in detail.get("标准地址", [])]
    duplicate_counts = pd.Series(logical_values).value_counts().to_dict()
    for _, source in detail.iterrows():
        page = source.get("页码", "")
        page = "" if pd.isna(page) else page
        title = _text(source.get("页标题", ""))
        logical = _tia_logical_address(source.get("标准地址", ""))
        data_type = _tia_data_type(source.get("类型", ""), logical)
        comment = _clean_comment(source.get("说明", ""))
        if logical not in included:
            status = "未进入TIA主表：地址或数据类型无效"
        elif duplicate_counts.get(logical, 0) > 1:
            status = "同地址候选：TIA主表仅保留置信度和说明质量最高的一条"
        else:
            status = "已进入TIA主表"
        note_parts = [project]
        if page != "":
            note_parts.append(f"P{page}")
        if title:
            note_parts.append(title)
        if comment:
            note_parts.append(comment)
        note_parts.append(status)
        rows.append({
            "项目": project,
            "方向": _text(source.get("类型", "")),
            "地址": _text(source.get("标准地址", "")),
            "中文说明": comment,
            "数据类型": data_type,
            "页码": page,
            "页标题": title,
            "源文件": source_name,
            "注释": " ".join(note_parts),
        })
    worksheet = workbook.create_sheet("原始提取_含重复")
    _append_frame(worksheet, pd.DataFrame(rows, columns=[
        "项目", "方向", "地址", "中文说明", "数据类型", "页码", "页标题", "源文件", "注释"
    ]))
    for index, width in enumerate([10, 10, 14, 40, 12, 8, 24, 30, 48], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    _style_header(worksheet, RAW_BROWN, 9)
    worksheet.freeze_panes = "A2"
    for row in worksheet.iter_rows(min_row=2, max_col=9):
        for cell in row:
            cell.font = Font(name="Carlito", size=11)
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column in {4, 7, 8, 9})


def _make_steps_sheet(workbook: Workbook, import_sheet: str) -> None:
    worksheet = workbook.create_sheet("TIA导入步骤")
    steps = [
        (1, "打开对应 PLC 的 TIA Portal 项目，并先创建项目备份。"),
        (2, f"打开工作表“{import_sheet}”，重点复核 Name、Data Type、Logical Address 和 Comment。"),
        (3, "在 PLC 变量表中执行“从文件导入”，选择此 XLSX；若当前 TIA 版本要求其自身模板，请先从 TIA 导出空表，再粘贴前 8 列。"),
        (4, "项目较多时按 TIA_<项目> 工作表分别导入，避免导入到错误的 PLC。"),
        (5, "导入后检查重复名称、地址冲突、安全点位和模拟量数据类型，再下载到 PLC。"),
    ]
    worksheet.append(["步骤", "操作"])
    for step in steps:
        worksheet.append(step)
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 90
    _style_header(worksheet, STEP_GREEN, 2)
    for row in worksheet.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.font = Font(name="Carlito", size=11)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_tia_xlsx(detail: pd.DataFrame, output_path: Path, source_path: Path | None = None) -> Path:
    """Export a reference-style TIA workbook with audit and instruction sheets."""
    suffix = "" if "TIA可导入" in output_path.stem else "_TIA"
    path = output_path.with_name(f"{output_path.stem}{suffix}.xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    project = infer_project_name(source_path)
    tia_frame = build_tia_table(detail, source_path, project)
    project_sheet = f"TIA_{project}"[:31]
    workbook = Workbook()
    workbook._named_styles["Normal"].font = Font(name="Carlito", size=11)
    _make_summary_sheet(workbook, tia_frame, project, project_sheet)
    _make_tia_sheet(workbook, "TIA_All_总表", tia_frame, "Table_TIA_All")
    _make_tia_sheet(workbook, project_sheet, tia_frame, "Table_TIA_Project")
    _make_raw_sheet(workbook, detail, tia_frame, project, source_path.name if source_path else "")
    _make_steps_sheet(workbook, project_sheet)
    workbook.save(path)
    return path

