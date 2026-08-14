from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from eplan_tag_exporter.exporters import (
    REFERENCE_COLUMNS,
    TIA_COLUMNS,
    build_tia_table,
    export_tia_csv,
    export_tia_xlsx,
)
from eplan_tag_exporter.io_service import export_outputs


def sample_detail() -> pd.DataFrame:
    return pd.DataFrame([
        {"页码": 11, "标准地址": "I0.0", "类型": "DI", "说明": "启动按钮"},
        {"页码": 11, "标准地址": "I0.0", "类型": "DI", "说明": "启动按钮副本"},
        {"页码": 12, "标准地址": "Q0.1", "类型": "DO", "说明": "运行灯"},
        {"页码": 13, "标准地址": "IW64", "类型": "AI", "说明": "压力值"},
        {"页码": 14, "标准地址": "DB1.DBX0.0", "类型": "DB", "说明": "状态"},
    ])


def test_build_tia_table_matches_reference_columns() -> None:
    frame = build_tia_table(sample_detail(), Path("PK06_PLC变量表.xlsx"))
    assert list(frame.columns) == REFERENCE_COLUMNS
    assert frame.loc[0, "Name"] == "PK06_I0_0"
    assert len(frame) == 4
    assert frame.loc[0, "Comment"] == "启动按钮"
    assert frame.loc[0, "Logical Address"] == "%I0.0"
    assert frame.loc[1, "Logical Address"] == "%IW64"
    assert frame.loc[1, "Data Type"] == "Int"
    assert frame.loc[2, "Logical Address"] == "%Q0.1"
    assert frame.loc[3, "Data Type"] == "Bool"
    assert frame.loc[0, "Hmi Writable"] == "True"


def test_tia_csv_contains_eight_import_columns(tmp_path: Path) -> None:
    output = export_tia_csv(sample_detail(), tmp_path / "output.xlsx", Path("PK06.pdf"))
    frame = pd.read_csv(output, encoding="utf-8-sig")
    assert list(frame.columns) == TIA_COLUMNS
    assert frame.loc[0, "Name"] == "PK06_I0_0"
    assert frame.loc[0, "Logical Address"] == "%I0.0"


def test_tia_xlsx_matches_reference_sheet_and_style_structure(tmp_path: Path) -> None:
    output = export_tia_xlsx(sample_detail(), tmp_path / "output.xlsx", Path("PK06_套后精平图.pdf"))
    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "说明_汇总", "TIA_All_总表", "TIA_PK06", "原始提取_含重复", "TIA导入步骤"
    ]
    all_tags = workbook["TIA_All_总表"]
    assert [cell.value for cell in all_tags[1]] == REFERENCE_COLUMNS
    assert all_tags["A2"].value == "PK06_I0_0"
    assert all_tags["D2"].value == "%I0.0"
    assert all_tags.tables["Table_TIA_All"].tableStyleInfo.name == "TableStyleMedium2"
    assert all_tags["A1"].fill.fgColor.rgb == "002F5597"
    assert all_tags.column_dimensions["E"].width == 42
    assert "A1:H1" in {str(item) for item in workbook["说明_汇总"].merged_cells.ranges}
    assert workbook["说明_汇总"]["B5"].value == 1
    assert workbook["说明_汇总"]["C5"].value == 1
    assert workbook["原始提取_含重复"]["A1"].fill.fgColor.rgb == "007F6000"
    assert workbook["TIA导入步骤"]["A1"].fill.fgColor.rgb == "00548235"
    assert all_tags.freeze_panes == "A2"


def test_tia_table_filters_invalid_rows_and_naturally_sorts_addresses() -> None:
    detail = pd.DataFrame([
        {"标准地址": "Q10.0", "类型": "DO", "说明": "输出十"},
        {"标准地址": "ABC", "类型": "Unknown", "说明": "无效"},
        {"标准地址": "I2.0", "类型": "DI", "说明": "输入二"},
        {"标准地址": "I10.0", "类型": "DI", "说明": "输入十"},
        {"标准地址": "I1.0", "类型": "DI", "说明": "输入一"},
    ])
    frame = build_tia_table(detail, Path("PK06.pdf"))
    assert frame["Logical Address"].tolist() == ["%I1.0", "%I2.0", "%I10.0", "%Q10.0"]


def test_comment_cleanup_prefers_description_over_neighboring_noise() -> None:
    detail = pd.DataFrame([
        {"标准地址": "I0.0", "类型": "DI", "说明": "CPU-DI1-1 X1.1 | 启动按钮"},
    ])
    frame = build_tia_table(detail, Path("PK06.pdf"))
    assert frame.loc[0, "Comment"] == "启动按钮"


def test_regular_and_tia_excel_do_not_overwrite_each_other(tmp_path: Path) -> None:
    source = tmp_path / "PK06.csv"
    source.write_text("名称,地址,说明\nStart,I0.0,启动\n", encoding="utf-8-sig")
    result = export_outputs(
        source,
        tmp_path / "PK06_PLC变量表_TIA可导入.xlsx",
        plc_vendor="siemens",
        formats=("xlsx", "tia_xlsx"),
    )
    assert len(set(result.files)) == 2
    assert all(path.exists() for path in result.files)
    assert any(path.stem.endswith("_识别明细") for path in result.files)

